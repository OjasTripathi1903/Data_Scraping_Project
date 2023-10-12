import pandas as pd
import time
import openpyxl
import extractor as ex

sheets = ["2022", "2021", "2020", "2019"]


# Check if the sheet already exists inside the book
def sheet_exists(year, sheets: list[str]):
    output = False
    for years in sheets:
        if year in years:
            output = True
            return output
        else:
            continue
    return output


def year_splitter(year: str, series: pd.Series):
    sheetname = year
    temp = series.iloc[0]
    print(temp)
    if year in str(temp):
        series.iloc[0] = year
    return pd.DataFrame(series).T


def split(raw_frame: pd.DataFrame, symbol: str):
    #    TODO: Loop this
    t2022 = pd.Series()
    t2021 = pd.Series()
    t2020 = pd.Series()
    t2019 = pd.Series()

    book = openpyxl.load_workbook("Data.xlsx")

    sheetnames = book.sheetnames
    print(sheetnames)
    indexes = list(raw_frame.iloc[0:, 0])
    for row in range(len(indexes)):
        check = str(indexes[row])

        if "2022" in check:
            print(raw_frame.iloc[row, :])
            output = raw_frame.iloc[row, :]
            output.iloc[0] = symbol
            output = pd.DataFrame(output).T

            ex.excel_inserter(
                indexing=False,
                dataframe=output,
                file_name="output.xlsx",
                sheetname="2022",
                append_by="row",
            )
        if "2021" in check:
            print(raw_frame.iloc[row, :])
            output = raw_frame.iloc[row, :]
            output.iloc[0] = symbol
            output = pd.DataFrame(output).T

            ex.excel_inserter(
                indexing=False,
                dataframe=output,
                file_name="output.xlsx",
                sheetname="2021",
                append_by="row",
            )
        if "2020" in check:
            print(raw_frame.iloc[row, :])
            output = raw_frame.iloc[row, :]
            output.iloc[0] = symbol
            output = pd.DataFrame(output).T

            ex.excel_inserter(
                indexing=False,
                dataframe=output,
                file_name="output.xlsx",
                sheetname="2020",
                append_by="row",
            )
        if "2019" in check:
            print(raw_frame.iloc[row, :])
            output = raw_frame.iloc[row, :]
            output.iloc[0] = symbol
            output = pd.DataFrame(output).T

            ex.excel_inserter(
                indexing=False,
                dataframe=output,
                file_name="output.xlsx",
                sheetname="2019",
                append_by="row",
            )
        else:
            print("Nope")
    return


def runner():
    masterframe = pd.read_excel("Data.xlsx", sheet_name="Master")
    test_series = masterframe.iloc[0]
    test_series = year_splitter("2022", test_series)
    # split(masterframe, "MMM", output_names)

    # ex.excel_inserter(test_series, "output.xlsx", "2022", "row", False)
    split(masterframe, "MMM")


if __name__ == "__main__":
    print(runner())
