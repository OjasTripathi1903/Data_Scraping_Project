import yfinance as yf
import pandas as pd
import prettify as pfy
import openpyxl as xl


def get_employee_count(symbol):
    logo = yf.Ticker(symbol)
    output = logo.info

    filtered_data = {"Full Time Employees": [output.get("fullTimeEmployees")]}
    employee_df = pd.DataFrame.from_dict(filtered_data, orient="columns")

    return employee_df


def get_income_statement(symbol):
    # Use the yfinance package to retrieve the income statement data
    stock = yf.Ticker(symbol)

    # Get the income statement
    income_statement = stock.financials

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(income_statement)
    # Only output the specific values
    output = pfy.prettifyIncome(df)
    return output


def get_balance_statement(symbol):
    # Use the yfinance package to retrieve the income statement data
    stock = yf.Ticker(symbol)

    # Get the balance statement
    balancesheet = stock.balance_sheet

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(balancesheet)

    output = pfy.prettifyBalance(df)

    return output


def get_cash_flow(symbol):
    # Use the yfinance package to retrieve the income statement data
    stock = yf.Ticker(symbol)
    # Get the cash flow
    cash_flow = stock.cashflow
    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(cash_flow)
    output = pfy.prettifyCash(df)
    return output


def excel_inserter(
    indexing: bool,
    dataframe: pd.DataFrame,
    file_name: str,
    sheetname: str,
    append_by: str,
):
    # Save the DataFrame to an Excel sheet

    try:
        # Try to load the existing workbook
        workbook = xl.load_workbook(file_name)
        try:
            sheet = workbook[sheetname]  # Get the active sheet
        except KeyError:
            sheet = workbook.create_sheet(sheetname)

    except FileNotFoundError:
        workbook = xl.Workbook()
        sheet = workbook.create_sheet(sheetname)
        workbook.save(file_name)
    location_row = sheet.max_row
    location_col = sheet.max_column

    with pd.ExcelWriter(
        path=file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay"
    ) as writer:
        if append_by == "row":
            dataframe.to_excel(
                writer,
                sheet_name=sheetname,
                index=True,
                startrow=location_row,
                header=indexing,
            )
        if append_by == "column":
            dataframe.to_excel(
                writer,
                sheet_name=sheetname,
                index=True,
                startcol=location_col,
                header=indexing,
            )
        else:
            dataframe.to_excel(writer, sheet_name=sheetname, index=True, header=True)


def spacer_creater(index_title: str):
    data = {index_title: ""}
    output = pd.DataFrame(list(data.values()), index=data.keys())
    return output


logo = "AOS"


def master_output(symbol):
    # Define the stock symbol for NVIDIA Corporation

    # Make API request to get the income statement data
    employee_count = get_employee_count(symbol)
    income_statement = get_income_statement(symbol)
    balance_sheet = get_balance_statement(symbol)
    cash_flow = get_cash_flow(symbol)

    # Create Spacers for neatness

    # TODO: Loop

    spacer_income_statement = spacer_creater("Income Statement")
    spacer_balance_sheet = spacer_creater("Balance Sheet")
    spacer_cash_flow = spacer_creater("Cash Flow")
    spacer_employee_count = spacer_creater("Employee Count")

    array = [income_statement, balance_sheet, cash_flow]
    # Save the data to an Excel sheet
    combined = pd.concat(array, axis=1, join="outer")
    excel_inserter(
        indexing=True,
        dataframe=combined,
        file_name="Data.xlsx",
        sheetname="Master",
        append_by="row",
    )
    excel_inserter(
        indexing=True,
        dataframe=employee_count,
        file_name="Data.xlsx",
        sheetname="Master",
        append_by="column",
    )
    return combined


if __name__ == "__main__":
    master_output(logo)
